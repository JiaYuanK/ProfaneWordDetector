import openpyxl as xl
import re 

#Edit distance algorithm
def editdistance(String1, String2):
    diff =  len(String1)-len(String2)
    
    diffnum = abs(diff)

    if len(String1) <= len(String2):
        for i in range(len(String1)):
            if String1[i] != String2[i]:
                diffnum += 1
    elif len(String2) < len(String1):
        for i in range(len(String2)):
            if String2[i] != String1[i]:
                diffnum += 1

    return diffnum

#Main program
def main():

    #Loads workbook
    workbook = xl.load_workbook(filename="badword-wmatrixV2.xlsx")

    ws = workbook.active

    smallestdistance = 100

    position = 0

    foundnum = []
    foundpos = []
    
    FUInput = input("Please input your profane word\n")

    if FUInput == "":
        print ("No word detected")
    else:
        #Changes all words to lowercase and removes the spaces between them
        FUNInput = FUInput.lower().replace(" ","")
        if "*" in FUNInput:
            UInput = FUNInput
        else:
            UInput = re.sub(r'(.)\1{2,}', r'\1',FUNInput )
        
        #Finds the smallest edit distance
        for row in ws.rows:
            if row[0].value.startswith(UInput[0]):
                EditDis = editdistance(UInput,row[0].value)
                foundnum.append(EditDis)

                if EditDis < smallestdistance:
                    smallestdistance = EditDis
                    
        if not foundnum:
            print("no words detected")
        else:
            #Finds the profane words with the lowest edit distance
            smallestdistance = min(foundnum)  
            for row in ws.rows:
                if row[0].value.startswith(UInput[0]):
                    EditDis = editdistance(UInput,row[0].value)
                    if EditDis == smallestdistance:
                        position = row[6].value
                        foundpos.append(position)

        #Printing results
        for i in foundpos:
            #Profane Word
            print("The profane word detected is",ws.cell(row=(i+1),column=1).value)
            #Category
            print("The category of the word is",ws.cell(row=(i+1),column=5).value)

    

main()

